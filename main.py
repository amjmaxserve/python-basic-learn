"""
Different ways to work with file

python has several built in functions for handling files in general. 

IO module: Create, read, write

python packages to work spreadsheet specifically: Openpyxl

Openpyxl is a third-party library for reading and writing Excel 

"""

import openpyxl


inv_file = openpyxl.load_workbook('inventory.xlsx')
product_list = inv_file["Inventory"]

products_per_supplier = {}
total_inventory_per_supplier = {}
total_value_per_supplier = {}

for product_row in range(2, product_list.max_row +1):
    supplier_name = product_list.cell(product_row, 4).value

    
    # calculation number of products per supplier
    if supplier_name in products_per_supplier:
        current_num_products = products_per_supplier.get(supplier_name)
        products_per_supplier[supplier_name] = current_num_products + 1

    else:
#        print("Adding a new supplier...")
        products_per_supplier[supplier_name] = 1


    
# Calculation of total invetory per supplier
for product_row in range(2, product_list.max_row +1):
    supplier_name = product_list.cell(product_row, 4).value
    inventory = product_list.cell(product_row, 2).value
    price = product_list.cell(product_row, 3).value
    
    
    if supplier_name in total_inventory_per_supplier:
        current_total_inventory = total_inventory_per_supplier.get(supplier_name)
        total_inventory_per_supplier[supplier_name] = current_total_inventory + inventory
    else:
 #       print("Adding a new supplier...")
        total_inventory_per_supplier[supplier_name] = inventory
        

# Calculation of total value of inventory per supplier
for product_row in range(2, product_list.max_row + 1):
    supplier_name = product_list.cell(product_row, 4).value
    inventory = product_list.cell(product_row, 2).value
    price = product_list.cell(product_row, 3).value

    if supplier_name in total_value_per_supplier:
        current_total_value = total_value_per_supplier.get(supplier_name)
        total_value_per_supplier[supplier_name] = current_total_value + (inventory * price)
    else:
        total_value_per_supplier[supplier_name] = inventory * price


from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ── Write "Total Inventory Price" as a new column 5 in the Excel sheet ────────

# Header styling (matches existing header style)
HEADER_FILL  = PatternFill("solid", fgColor="1E3A5F")
HEADER_FONT  = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center")
thin         = Side(style="thin", color="BFBFBF")
BORDER       = Border(left=thin, right=thin, top=thin, bottom=thin)
DATA_FONT    = Font(name="Calibri", size=11)
DATA_ALIGN   = Alignment(horizontal="center", vertical="center")

# Write header
header_cell = product_list.cell(1, 5, "Total Inventory Price ($)")
header_cell.font      = HEADER_FONT
header_cell.fill      = HEADER_FILL
header_cell.alignment = HEADER_ALIGN
header_cell.border    = BORDER
product_list.column_dimensions["E"].width = 28

# Write per-row total value for the row's supplier
for product_row in range(2, product_list.max_row + 1):
    supplier_name = product_list.cell(product_row, 4).value
    total_value   = round(total_value_per_supplier.get(supplier_name, 0), 2)

    cell = product_list.cell(product_row, 5, total_value)
    cell.font      = DATA_FONT
    cell.alignment = DATA_ALIGN
    cell.border    = BORDER

# Save as a new file (keeps the original inventory.xlsx untouched)
inv_file.save("inventory_with_total_value.xlsx")
print("✅  inventory_with_total_value.xlsx created with 'Total Inventory Price' column.")

# Summary printout
print("\n── Products per supplier ──")
for supplier, count in sorted(products_per_supplier.items()):
    print(f"  {supplier:<30} {count} products")

print("\n── Total inventory units per supplier ──")
for supplier, total in sorted(total_inventory_per_supplier.items()):
    print(f"  {supplier:<30} {total} units")

print("\n── Total inventory value per supplier ──")
for supplier, value in sorted(total_value_per_supplier.items()):
    print(f"  {supplier:<30} ${value:,.2f}")
