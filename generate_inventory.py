import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import random

# ── Supplier pool ────────────────────────────────────────────────────────────
SUPPLIERS = [
    "TechSupplies Co.",
    "Global Parts Ltd.",
    "FastTrack Wholesale",
    "Prime Distributors",
    "Apex Trade Group",
    "NextGen Supply",
    "BlueStar Logistics",
    "EcoSource Inc.",
    "Pinnacle Traders",
    "Vertex Supply Chain",
]

# ── Category → base price range ──────────────────────────────────────────────
CATEGORIES = {
    "ELEC": (15.00, 999.99),
    "FURN": (50.00, 1500.00),
    "CLTH": (5.00, 250.00),
    "FOOD": (1.00, 80.00),
    "TOOL": (10.00, 500.00),
    "BOOK": (3.00, 75.00),
    "SPRT": (8.00, 600.00),
    "HOME": (5.00, 400.00),
}

# ── Build workbook ────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Inventory"

# ── Header styling ────────────────────────────────────────────────────────────
HEADER_FILL   = PatternFill("solid", fgColor="1E3A5F")       # deep navy
HEADER_FONT   = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
HEADER_ALIGN  = Alignment(horizontal="center", vertical="center")

thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

HEADERS = ["Product Number", "Inventory", "Price ($)", "Supplier"]
COL_WIDTHS = [20, 14, 14, 28]

for col_idx, (header, width) in enumerate(zip(HEADERS, COL_WIDTHS), start=1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.font   = HEADER_FONT
    cell.fill   = HEADER_FILL
    cell.alignment = HEADER_ALIGN
    cell.border = BORDER
    ws.column_dimensions[get_column_letter(col_idx)].width = width

ws.row_dimensions[1].height = 28

# ── Alternate row fills ───────────────────────────────────────────────────────
FILL_ODD  = PatternFill("solid", fgColor="F0F4FA")
FILL_EVEN = PatternFill("solid", fgColor="FFFFFF")
DATA_FONT  = Font(name="Calibri", size=11)
DATA_ALIGN = Alignment(horizontal="center", vertical="center")

# ── Generate 1 000 rows ───────────────────────────────────────────────────────
random.seed(42)          # reproducible data


for row_idx in range(2, 1002):
    # Sequential product number: 1 to 1000
    product_number = row_idx - 1

    inventory = random.randint(0, 2000)
    category  = random.choice(list(CATEGORIES.keys()))
    low, high = CATEGORIES[category]
    price     = round(random.uniform(low, high), 2)
    supplier  = random.choice(SUPPLIERS)

    row_data = [product_number, inventory, price, supplier]
    fill = FILL_ODD if (row_idx % 2 == 0) else FILL_EVEN

    for col_idx, value in enumerate(row_data, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.font      = DATA_FONT
        cell.fill      = fill
        cell.alignment = DATA_ALIGN
        cell.border    = BORDER

    # Colour inventory cell red when stock is low (< 20)
    if inventory < 20:
        ws.cell(row=row_idx, column=2).font = Font(
            name="Calibri", size=11, bold=True, color="C0392B"
        )

# ── Freeze top row, add auto-filter ──────────────────────────────────────────
ws.freeze_panes = "A2"
ws.auto_filter.ref = ws.dimensions

# ── Save ─────────────────────────────────────────────────────────────────────
output_path = "inventory.xlsx"
wb.save(output_path)
print(f"✅  inventory.xlsx created — {ws.max_row - 1} data rows written.")
